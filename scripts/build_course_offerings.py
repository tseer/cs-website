#!/usr/bin/env python3
from __future__ import annotations

from collections import OrderedDict
from copy import deepcopy
from datetime import UTC, date, datetime
from pathlib import Path
import os
import re
from typing import Any

import yaml
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = REPO_ROOT / "_data" / "course_offerings.yml"
PEOPLE_DIR = REPO_ROOT / "_people"
COURSES_DIR = REPO_ROOT / "WEB" / "academics" / "courses"
FILE_PATTERN = re.compile(r"^schedule_(fall|spring|summer(?:_session[123])?)_(\d{4})\.xlsx$")

TERM_CONFIG = {
    "fall": {"label": "Fall", "order": 1},
    "spring": {"label": "Spring", "order": 2},
    "summer_session1": {"label": "Summer Session 1", "order": 3},
    "summer_session2": {"label": "Summer Session 2", "order": 4},
    "summer_session3": {"label": "Summer Session 3", "order": 5},
    "summer": {"label": "Summer", "order": 3},
}

HEADER_ALIASES = {
    "subject": "subject",
    "ctlg": "catalog_number",
    "sect": "section",
    "title": "title",
    "cls": "class_number",
    "cmpt": "component",
    "component": "component",
    "facid": "location",
    "room": "location",
    "location": "location",
    "mtgstart": "meeting_start",
    "mtgend": "meeting_end",
    "mtgptrn": "meeting_pattern",
    "empname": "instructor",
    "fullname": "instructor",
    "instructor": "instructor",
    "mtgptrnstartdt": "meeting_pattern_start_date",
    "mtgptrnenddt": "meeting_pattern_end_date",
    "first": "first_name",
    "last": "last_name",
}


def normalize_header(value: Any) -> str:
    text = stringify(value)
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def normalize_section(value: Any) -> str:
    text = stringify(value)
    if not text:
        return ""
    if text.isdigit() and len(text) < 2:
        return text.zfill(2)
    return text


def normalize_catalog_number(value: Any) -> str:
    return stringify(value).replace(" ", "")


def format_instructor(value: str, first_name: str = "", last_name: str = "") -> str:
    text = stringify(value)
    if not text and (first_name or last_name):
        text = " ".join(part for part in [stringify(first_name), stringify(last_name)] if part)
    text = re.sub(r"\s+", " ", text)
    if "," in text:
        parts = [part.strip() for part in text.split(",") if part.strip()]
        if len(parts) >= 2:
            text = ", ".join(parts)
    return text


def normalize_meeting_line(days: str, start: str, end: str) -> str:
    days = stringify(days)
    start = stringify(start)
    end = stringify(end)
    if days == "TBA" and not start and not end:
        return "TBA"
    if start and end:
        time_part = f"{start} - {end}"
    else:
        time_part = start or end
    if days and time_part:
        return f"{days} {time_part}"
    return days or time_part or "TBA"


def normalize_name_text(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[.,]+", " ", stringify(value).lower())).strip()


def name_tokens(value: str) -> list[str]:
    return [token for token in normalize_name_text(value).split(" ") if token]


def split_instructor_names(value: str) -> list[str]:
    text = stringify(value)
    if not text:
        return []
    normalized = (
        text.replace(" / ", " ; ")
        .replace(" & ", " ; ")
        .replace(" and ", " ; ")
        .replace("; ", ";")
        .replace(" ;", ";")
    )
    return [part.strip() for part in normalized.split(";") if part.strip()]


def load_front_matter(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8")
    if not text.startswith("---"):
        return {}
    parts = text.split("---", 2)
    if len(parts) < 3:
        return {}
    return yaml.safe_load(parts[1]) or {}


def build_people_index() -> dict[str, dict[str, Any]]:
    people_index: dict[str, dict[str, Any]] = {}
    for path in sorted(PEOPLE_DIR.glob("*.md")):
        data = load_front_matter(path)
        slug = path.stem
        person_name = stringify(data.get("person_name") or data.get("title"))
        aliases = [stringify(alias) for alias in (data.get("aliases") or []) if stringify(alias)]
        variants = []
        for candidate in [person_name, stringify(data.get("title"))] + aliases:
            if candidate and candidate not in variants:
                variants.append(candidate)
        people_index[slug] = {
            "slug": slug,
            "person_name": person_name,
            "url": stringify(data.get("permalink")),
            "aliases": aliases,
            "variants": variants,
        }
    return people_index


def build_course_page_index() -> dict[str, dict[str, Any]]:
    course_index: dict[str, dict[str, Any]] = {}
    for path in sorted(COURSES_DIR.glob("*.markdown")):
        data = load_front_matter(path)
        course_code = stringify(data.get("course_code"))
        if not course_code:
            continue
        course_index[course_code] = {
            "course_code": course_code,
            "course_name": stringify(data.get("course_name")),
            "title": stringify(data.get("title")),
            "url": stringify(data.get("permalink")),
            "keywords": [stringify(keyword) for keyword in (data.get("keywords") or []) if stringify(keyword)],
        }
    return course_index


def match_person(name: str, people_index: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    normalized = normalize_name_text(name)
    if not normalized:
        return None

    instructor_token_set = set(name_tokens(name))
    best_match = None
    best_score = -1
    tied = False

    for person in people_index.values():
        person_best_score = -1
        for variant in person["variants"]:
            variant_normalized = normalize_name_text(variant)
            variant_token_set = set(name_tokens(variant))

            if variant_normalized == normalized:
                person_best_score = max(person_best_score, 3)
            elif instructor_token_set and variant_token_set and instructor_token_set == variant_token_set and len(variant_token_set) >= 2:
                person_best_score = max(person_best_score, 2)
            elif instructor_token_set and variant_token_set and len(variant_token_set) >= 2:
                if variant_token_set.issubset(instructor_token_set) or instructor_token_set.issubset(variant_token_set):
                    person_best_score = max(person_best_score, 1)

        if person_best_score > best_score:
            best_match = person
            best_score = person_best_score
            tied = False
        elif person_best_score == best_score and person_best_score >= 0:
            tied = True

    if tied or best_score < 0:
        return None
    return best_match


def build_relationships(
    academic_years: list[dict[str, Any]],
    people_index: dict[str, dict[str, Any]],
    course_page_index: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    people_relationships: dict[str, dict[str, Any]] = {
        slug: {
            "related_courses": [],
            "related_course_entries": [],
            "keyword_supplements": [],
        }
        for slug in people_index.keys()
    }
    course_relationships: dict[str, dict[str, Any]] = {}
    instructor_matches: dict[str, dict[str, Any]] = {}

    for academic_year in academic_years:
        for term in academic_year["terms"]:
            for offering in term["offerings"]:
                course_code = offering["course_code"]
                course_entry = course_relationships.setdefault(
                    course_code,
                    {
                        "related_people": [],
                        "related_people_entries": [],
                        "keyword_supplements": [],
                    },
                )

                for instructor_name in split_instructor_names(offering.get("instructor", "")):
                    matched_person = match_person(instructor_name, people_index)
                    if not matched_person:
                        continue

                    instructor_matches[instructor_name] = {
                        "slug": matched_person["slug"],
                        "person_name": matched_person["person_name"],
                        "url": matched_person["url"],
                    }

                    if matched_person["slug"] not in course_entry["related_people"]:
                        course_entry["related_people"].append(matched_person["slug"])
                        course_entry["related_people_entries"].append(
                            {
                                "slug": matched_person["slug"],
                                "person_name": matched_person["person_name"],
                                "url": matched_person["url"],
                            }
                        )

                    if matched_person["person_name"] not in course_entry["keyword_supplements"]:
                        course_entry["keyword_supplements"].append(matched_person["person_name"])

                    if course_code in course_page_index:
                        person_entry = people_relationships[matched_person["slug"]]
                        if course_code not in person_entry["related_courses"]:
                            person_entry["related_courses"].append(course_code)
                            course_page = course_page_index[course_code]
                            person_entry["related_course_entries"].append(
                                {
                                    "course_code": course_code,
                                    "course_name": course_page.get("course_name") or course_page.get("title"),
                                    "url": course_page.get("url"),
                                }
                            )
                        for keyword in [course_code, course_page_index[course_code].get("course_name") or course_page_index[course_code].get("title")]:
                            keyword = stringify(keyword)
                            if keyword and keyword not in person_entry["keyword_supplements"]:
                                person_entry["keyword_supplements"].append(keyword)

    return {
        "people": people_relationships,
        "courses": course_relationships,
        "instructor_matches": instructor_matches,
    }


def current_academic_year_start(today: date) -> int:
    return today.year if today.month >= 8 else today.year - 1


def academic_year_start(term_type: str, year: int) -> int:
    if term_type == "fall":
        return year
    return year - 1


def academic_year_label(start_year: int) -> str:
    return f"{start_year}-{start_year + 1}"


def term_label(term_type: str, year: int) -> str:
    return f"{TERM_CONFIG[term_type]['label']} {year}"


def parse_today() -> date:
    override = os.environ.get("COURSE_OFFERINGS_TODAY")
    if override:
        return date.fromisoformat(override)
    return date.today()


def discover_schedule_files(root: Path) -> list[tuple[Path, str, int]]:
    matches: list[tuple[Path, str, int]] = []
    for path in sorted(root.glob("schedule_*.xlsx")):
        match = FILE_PATTERN.match(path.name)
        if not match:
            continue
        raw_term, year_text = match.groups()
        matches.append((path, raw_term, int(year_text)))
    return matches


def load_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration:
        return []

    normalized_headers = [HEADER_ALIASES.get(normalize_header(value), normalize_header(value)) for value in headers]
    rows: list[dict[str, Any]] = []
    for row in iterator:
        record = {}
        for index, header in enumerate(normalized_headers):
            if not header:
                continue
            record[header] = row[index] if index < len(row) else None
        rows.append(record)
    return rows


def build_offerings_for_term(rows: list[dict[str, Any]], *, term_type: str, year: int, source_file: str) -> list[dict[str, Any]]:
    aggregated: "OrderedDict[tuple[str, str, str, str, str], dict[str, Any]]" = OrderedDict()

    for row in rows:
        subject = stringify(row.get("subject")).upper()
        catalog_number = normalize_catalog_number(row.get("catalog_number"))
        title = stringify(row.get("title"))
        section = normalize_section(row.get("section"))

        if not subject or not catalog_number or not title or not section:
            continue

        course_code = f"{subject}{catalog_number}"
        class_number = stringify(row.get("class_number"))
        component = stringify(row.get("component"))
        location = stringify(row.get("location"))
        instructor = format_instructor(
            stringify(row.get("instructor")),
            first_name=stringify(row.get("first_name")),
            last_name=stringify(row.get("last_name")),
        )

        meeting = {
            "days": stringify(row.get("meeting_pattern")),
            "start": stringify(row.get("meeting_start")),
            "end": stringify(row.get("meeting_end")),
            "location": location,
        }
        if row.get("meeting_pattern_start_date"):
            meeting["start_date"] = stringify(row.get("meeting_pattern_start_date"))
        if row.get("meeting_pattern_end_date"):
            meeting["end_date"] = stringify(row.get("meeting_pattern_end_date"))

        offering_key = (course_code, title, section, class_number, component)
        if offering_key not in aggregated:
            aggregated[offering_key] = {
                "course_code": course_code,
                "subject": subject,
                "catalog_number": catalog_number,
                "title": title,
                "section": section,
                "class_number": class_number,
                "component": component,
                "instructor": instructor,
                "location": location,
                "term_type": term_type,
                "term_label": term_label(term_type, year),
                "term_slug": f"{term_type}_{year}",
                "term_year": year,
                "academic_year": academic_year_label(academic_year_start(term_type, year)),
                "academic_year_start": academic_year_start(term_type, year),
                "source_file": source_file,
                "meetings": [],
            }

        meeting_key = tuple(meeting.get(key, "") for key in ["days", "start", "end", "location", "start_date", "end_date"])
        existing_meeting_keys = {
            tuple(existing.get(key, "") for key in ["days", "start", "end", "location", "start_date", "end_date"])
            for existing in aggregated[offering_key]["meetings"]
        }
        if meeting_key not in existing_meeting_keys and any(meeting.values()):
            aggregated[offering_key]["meetings"].append(meeting)

        if not aggregated[offering_key]["instructor"] and instructor:
            aggregated[offering_key]["instructor"] = instructor
        if not aggregated[offering_key]["location"] and location:
            aggregated[offering_key]["location"] = location

    offerings = list(aggregated.values())
    for offering in offerings:
        meetings = offering["meetings"]
        offering["schedule_lines"] = [
            normalize_meeting_line(meeting.get("days", ""), meeting.get("start", ""), meeting.get("end", ""))
            for meeting in meetings
        ] or ["TBA"]
        location_lines = []
        seen_locations = set()
        for meeting in meetings:
            meeting_location = stringify(meeting.get("location"))
            if meeting_location and meeting_location not in seen_locations:
                location_lines.append(meeting_location)
                seen_locations.add(meeting_location)
        if not location_lines and offering["location"]:
            location_lines.append(offering["location"])
        offering["location_lines"] = location_lines or ["TBA"]
        offering["schedule_summary"] = "; ".join(offering["schedule_lines"])

    offerings.sort(key=lambda item: (item["course_code"], item["section"], item["class_number"], item["component"]))
    return offerings


def build_payload(today: date) -> dict[str, Any]:
    current_ay_start = current_academic_year_start(today)
    schedule_files = discover_schedule_files(REPO_ROOT)
    people_index = build_people_index()
    course_page_index = build_course_page_index()

    academic_year_map: "OrderedDict[int, dict[str, Any]]" = OrderedDict()
    course_index: dict[str, list[dict[str, Any]]] = {}
    included_files: list[str] = []
    skipped_files: list[str] = []

    for path, raw_term, year in schedule_files:
        if raw_term not in TERM_CONFIG:
            skipped_files.append(path.name)
            continue

        ay_start = academic_year_start(raw_term, year)
        if ay_start < current_ay_start:
            continue

        try:
            rows = load_rows(path)
        except Exception:
            skipped_files.append(path.name)
            continue

        offerings = build_offerings_for_term(rows, term_type=raw_term, year=year, source_file=path.name)
        included_files.append(path.name)

        ay_entry = academic_year_map.setdefault(
            ay_start,
            {
                "label": academic_year_label(ay_start),
                "start_year": ay_start,
                "end_year": ay_start + 1,
                "terms": [],
            },
        )

        term_entry = {
            "slug": f"{raw_term}_{year}",
            "label": term_label(raw_term, year),
            "type": raw_term,
            "year": year,
            "order": TERM_CONFIG[raw_term]["order"],
            "source_file": path.name,
            "offering_count": len(offerings),
            "offerings": offerings,
        }
        ay_entry["terms"].append(term_entry)

        for offering in offerings:
            course_index.setdefault(offering["course_code"], []).append(deepcopy(offering))

    academic_years = sorted(academic_year_map.values(), key=lambda item: item["start_year"])
    for ay in academic_years:
        ay["terms"].sort(key=lambda term: (term["order"], term["year"], term["slug"]))

    sorted_course_index = {
        course_code: sorted(
            offerings,
            key=lambda item: (item["academic_year_start"], TERM_CONFIG[item["term_type"]]["order"], item["section"], item["class_number"]),
        )
        for course_code, offerings in sorted(course_index.items())
    }

    relationships = build_relationships(academic_years, people_index, course_page_index)

    return {
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "generated_for_date": today.isoformat(),
        "current_academic_year": academic_year_label(current_ay_start),
        "current_academic_year_start": current_ay_start,
        "term_order": [
            {"type": "fall", "label": "Fall", "order": 1},
            {"type": "spring", "label": "Spring", "order": 2},
            {"type": "summer_session1", "label": "Summer Session 1", "order": 3},
            {"type": "summer_session2", "label": "Summer Session 2", "order": 4},
            {"type": "summer_session3", "label": "Summer Session 3", "order": 5},
        ],
        "source_files": included_files,
        "skipped_files": skipped_files,
        "academic_years": academic_years,
        "course_index": sorted_course_index,
        "relationships": relationships,
    }


def main() -> None:
    payload = build_payload(parse_today())
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_PATH.open("w", encoding="utf-8") as handle:
        yaml.safe_dump(payload, handle, sort_keys=False, allow_unicode=True, width=120)


if __name__ == "__main__":
    main()
